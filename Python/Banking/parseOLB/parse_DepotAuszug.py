"""
Parse Degussa Bank depot statements (PDF) and export to Excel.
Extracts security information including ISIN, quantity, name, price, and value.
"""

import pdfplumber
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path
import re
import datetime
from typing import List, Dict, Optional, Tuple


CANONICAL_NAMES_BY_ISIN = {
    "DE0008474008": "DWS ESG INVESTA INHABER-ANTEILE LD",
    "DE0009769794": "DWS ESG TOP WORLD INHABER-ANTEILE",
    "DE0008490673": "DEGUSSA BK.UNIVERSAL-RENTENFONDS INHABER-ANTEILE",
    "LU0553164731": "DJE - ZINS + DIVIDENDE INHABER-ANTEILE PA EUR O.N.",
    "DE000A12BSB8": "FOKUS WOHNEN DEUTSCHLAND INHABER-ANTEILE",
    "DE000ANTE1A3": "ANTEA INVTAG MVK U.TGV - ANTEA INHABER-ANLAGEAKTIEN R",
    "LU0323578657": "FLOSSB.V.STORCH-MULT.OPPORT. INHABER-ANTEILE R O.N.",
}


# German month name → month number
_GERMAN_MONTHS = {
    "januar": 1, "februar": 2, "märz": 3, "april": 4,
    "mai": 5, "juni": 6, "juli": 7, "august": 8,
    "september": 9, "oktober": 10, "november": 11, "dezember": 12,
    "jan": 1, "feb": 2, "mär": 3, "apr": 4,
    "jun": 6, "jul": 7, "aug": 8, "sep": 9, "okt": 10, "nov": 11, "dez": 12,
}


def _extract_depot_date(full_text: str, pdf_path: str) -> Optional[datetime.date]:
    """Return the depot statement date from PDF text or filename."""
    # 1) "Depotbestand per DD. MonatName YYYY"  (German long form)
    m = re.search(
        r"Depotbestand\s+per\s+(\d{1,2})\.?\s+([A-Za-zäöü]+)\s+(\d{4})",
        full_text, re.IGNORECASE,
    )
    if m:
        day, month_str, year = int(m.group(1)), m.group(2).lower(), int(m.group(3))
        # strip trailing dot that pdfplumber sometimes attaches
        month_str = month_str.rstrip(".")
        month = _GERMAN_MONTHS.get(month_str)
        if month:
            try:
                return datetime.date(year, month, day)
            except ValueError:
                pass

    # 2) "Depotbestand per DD.MM.YYYY"  (numeric form)
    m = re.search(
        r"Depotbestand\s+per\s+(\d{1,2})\.(\d{1,2})\.(\d{4})",
        full_text, re.IGNORECASE,
    )
    if m:
        try:
            return datetime.date(int(m.group(3)), int(m.group(2)), int(m.group(1)))
        except ValueError:
            pass

    # 3) Date from filename: "..._YYYY-MM-DD_..."
    fname = Path(pdf_path).stem
    m = re.search(r"(\d{4})-(\d{2})-(\d{2})", fname)
    if m:
        try:
            return datetime.date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        except ValueError:
            pass

    return None


def extract_securities_from_pdf(pdf_path: str) -> Tuple[List[Dict], Dict]:
    """
    Extract securities data from Degussa Bank depot PDF statement.

    Args:
        pdf_path: Path to the PDF file

    Returns:
        Tuple of (securities_list, summary_data)
    """
    securities = []
    summary_data = {}
    all_text_pages: List[str] = []

    with pdfplumber.open(pdf_path) as pdf:
        # Extract from all pages
        for page_num, page in enumerate(pdf.pages):
            print(f"Processing page {page_num + 1}...")

            # Extract text from page
            text = page.extract_text() or ""
            all_text_pages.append(text)

            # Try to extract tables
            tables = page.extract_tables()
            if tables:
                for table in tables:
                    _process_table(table, securities, summary_data)

            # Fallback: parse text for security information
            _parse_text_for_securities(text, securities)

    full_text = "\n".join(all_text_pages)
    depot_date = _extract_depot_date(full_text, pdf_path)

    securities = _backfill_security_numbers(securities)
    # Stamp the date onto every security row
    for sec in securities:
        sec["datum"] = depot_date

    return securities, summary_data


def _process_table(table: List[List[str]], securities: List[Dict], summary_data: Dict) -> None:
    """Process extracted table data."""
    if not table or len(table) < 2:
        return
    
    headers = table[0]
    
    # Check if this is a summary table (Depotstruktur)
    if any('Wertpapierart' in str(h) for h in headers):
        _process_summary_table(table, summary_data)
    else:
        # Process securities table
        _process_securities_table(table, securities)


def _process_summary_table(table: List[List[str]], summary_data: Dict) -> None:
    """Extract portfolio summary by asset type."""
    for row in table[1:]:
        if len(row) >= 3 and row[0] and row[0].strip():
            asset_type = row[0].strip()
            if 'Gesamtkurswert' not in asset_type:
                try:
                    percentage = float(row[2].replace('%', '').strip()) if len(row) > 2 else 0
                    value = float(row[3].replace('EUR', '').replace('.', '').replace(',', '.').strip()) if len(row) > 3 else 0
                    summary_data[asset_type] = {
                        'percentage': percentage,
                        'value': value
                    }
                except (ValueError, IndexError):
                    pass


def _process_securities_table(table: List[List[str]], securities: List[Dict]) -> None:
    """Extract individual securities from table."""
    current_category = None
    
    for row in table[1:]:
        if not row or not any(row):
            continue
        
        # Detect category headers
        if len(row) == 1 or (len(row) > 0 and row[0] and not row[1]):
            if row[0] and row[0].strip() in ['Aktienfonds', 'Rentenwerte', 'Mischfonds', 'Immobilienfonds']:
                current_category = row[0].strip()
            continue
        
        # Try to parse security data
        security = _parse_security_row(row, current_category)
        if security:
            securities.append(security)


def _parse_german_number(value: str) -> float | None:
    """Parse German/EN formatted number strings to float."""
    if not value:
        return None

    s = value.strip().replace("\xa0", " ")
    s = re.sub(r"[^0-9,.-]", "", s)
    if not s:
        return None

    try:
        # German style: 1.234,56
        if "," in s and "." in s:
            s = s.replace(".", "").replace(",", ".")
        elif "," in s:
            s = s.replace(",", ".")
        elif "." in s and re.fullmatch(r"\d{1,3}(?:\.\d{3})+", s):
            # Thousands separators only: 4.937
            s = s.replace(".", "")
        return float(s)
    except ValueError:
        return None


def _extract_numbers_from_cell(cell: str) -> List[float]:
    """Extract numeric tokens from a text cell using German number format awareness."""
    if not cell:
        return []
    tokens = re.findall(r"\d{1,3}(?:\.\d{3})*(?:,\d+)?|\d+(?:,\d+)?", cell)
    values = []
    for token in tokens:
        parsed = _parse_german_number(token)
        if parsed is not None:
            values.append(parsed)
    return values


def _extract_quantity_from_text_fragment(fragment: str) -> float | None:
    """Extract quantity from a text fragment using common unit patterns."""
    if not fragment:
        return None

    patterns = [
        # Negative lookahead prevents matching dates like "30.Aug.2024"
        r"(?:St[üu]ck|STK|Stk|Anteile?)\s+([0-9][0-9\.,]*)(?!\s*\.\s*[A-Za-z])",
        r"([0-9][0-9\.,]*)\s*(?:St[üu]ck|STK|Stk|Anteile?)",
    ]
    for pattern in patterns:
        m = re.search(pattern, fragment, re.IGNORECASE)
        if m:
            q = _parse_german_number(m.group(1))
            if q is not None and q > 0:
                return q
    return None


def _extract_eur_numbers_from_fragment(fragment: str) -> List[float]:
    """Extract numbers tied to EUR in a text fragment.
    Handles both '<number> EUR' and 'EUR <number>' formats.
    """
    if not fragment:
        return []
    seen: List[float] = []
    # number before EUR: e.g. "189,90 EUR"
    for m in re.findall(r"([0-9][0-9\.,]*)\s*EUR", fragment, re.IGNORECASE):
        parsed = _parse_german_number(m)
        if parsed is not None:
            seen.append(parsed)
    # EUR before number: e.g. "EUR 4.937,40" or "EUR 6.662,42"
    for m in re.findall(r"\bEUR\s+([0-9][0-9\.,]*)", fragment, re.IGNORECASE):
        parsed = _parse_german_number(m)
        if parsed is not None and parsed not in seen:
            seen.append(parsed)
    return seen


def _normalize_name_by_isin(isin: str, extracted_name: str) -> str:
    """Return canonical name for known ISINs; otherwise keep extracted name."""
    if not isin:
        return extracted_name or ""
    return CANONICAL_NAMES_BY_ISIN.get(isin, extracted_name or "")


def _format_eur_value(value: float | int | None) -> str:
    """Format numeric value for log output; handle missing values gracefully."""
    if value is None:
        return "n/a"
    return f"{float(value):10.2f}"


def _is_missing_number(value: float | int | None) -> bool:
    """Treat None and 0 as missing parser output for position/price/value fields."""
    return value is None or (isinstance(value, (int, float)) and value == 0)


def _backfill_security_numbers(securities: List[Dict]) -> List[Dict]:
    """Backfill missing quantity/price/value fields by ISIN and arithmetic consistency."""
    if not securities:
        return securities

    # First pass: infer within each row.
    for sec in securities:
        q = sec.get("quantity")
        p = sec.get("price")
        v = sec.get("value")

        if _is_missing_number(v) and not _is_missing_number(q) and not _is_missing_number(p):
            sec["value"] = float(q) * float(p)
            v = sec.get("value")
        if _is_missing_number(p) and not _is_missing_number(q) and not _is_missing_number(v) and float(q) != 0:
            sec["price"] = float(v) / float(q)
            p = sec.get("price")
        if _is_missing_number(q) and not _is_missing_number(p) and not _is_missing_number(v) and float(p) != 0:
            sec["quantity"] = float(v) / float(p)

    # Second pass: fill from other rows with same ISIN.
    by_isin: Dict[str, List[Dict]] = {}
    for sec in securities:
        isin = (sec.get("isin") or "").strip()
        if isin:
            by_isin.setdefault(isin, []).append(sec)

    for isin, rows in by_isin.items():
        known_q = next((r.get("quantity") for r in rows if not _is_missing_number(r.get("quantity"))), None)
        known_p = next((r.get("price") for r in rows if not _is_missing_number(r.get("price"))), None)
        known_v = next((r.get("value") for r in rows if not _is_missing_number(r.get("value"))), None)

        for r in rows:
            if _is_missing_number(r.get("quantity")) and known_q is not None:
                r["quantity"] = known_q
            if _is_missing_number(r.get("price")) and known_p is not None:
                r["price"] = known_p
            if _is_missing_number(r.get("value")) and known_v is not None:
                r["value"] = known_v

            # Re-run arithmetic fill after ISIN backfill.
            q = r.get("quantity")
            p = r.get("price")
            v = r.get("value")
            if _is_missing_number(v) and not _is_missing_number(q) and not _is_missing_number(p):
                r["value"] = float(q) * float(p)
            if _is_missing_number(p) and not _is_missing_number(q) and not _is_missing_number(v) and float(q) != 0:
                r["price"] = float(v) / float(q)
            if _is_missing_number(q) and not _is_missing_number(p) and not _is_missing_number(v) and float(p) != 0:
                r["quantity"] = float(v) / float(p)

    return securities


def _parse_security_row(row: List[str], category: str) -> Dict | None:
    """Parse a single security row."""
    if len(row) < 4:
        return None
    
    try:
        isin = None
        isin_idx = -1
        for cell in row:
            if cell and re.match(r'^[A-Z]{2}[A-Z0-9]{9}[0-9]$', cell.strip()):
                isin = cell.strip()
                isin_idx = row.index(cell)
                break

        # Quantity can appear before or after ISIN depending on table layout.
        quantity = None
        quantity_candidates = []

        # 1) Prefer explicit quantity markers in any cell.
        for cell in row:
            if not cell:
                continue
            q = _extract_quantity_from_text_fragment(cell)
            if q is not None:
                quantity = q
                break

        # 2) If not found, try cells around ISIN (common broker layouts).
        if quantity is None:
            if isin_idx >= 0:
                candidate_cells = []
                for idx in [isin_idx - 2, isin_idx - 1, isin_idx + 1, isin_idx + 2]:
                    if 0 <= idx < len(row):
                        candidate_cells.append(row[idx])
            else:
                candidate_cells = row[:3]

            for cell in candidate_cells:
                if not cell:
                    continue
                quantity_candidates.extend(_extract_numbers_from_cell(cell))

        # 3) Last resort: allow any integer-like positive number in the row,
        # then later exclude values that look like price/value.
        if quantity is None and not quantity_candidates:
            for cell in row:
                if not cell:
                    continue
                quantity_candidates.extend(_extract_numbers_from_cell(cell))

        if quantity is None and quantity_candidates:
            integer_like = [q for q in quantity_candidates if q > 0 and abs(q - round(q)) < 1e-9]
            quantity = integer_like[0] if integer_like else quantity_candidates[0]

        # Collect numeric info per cell for price/value inference
        per_cell_numbers = []
        all_numbers = []
        for idx, cell in enumerate(row):
            if not cell:
                continue
            vals = _extract_numbers_from_cell(cell)
            if not vals:
                continue
            has_eur = "EUR" in cell.upper()
            for v in vals:
                per_cell_numbers.append((idx, v, has_eur))
                all_numbers.append(v)

        price = None
        value = None

        eur_values = [v for idx, v, has_eur in per_cell_numbers if has_eur]
        if eur_values:
            price = eur_values[0]
            if len(eur_values) > 1:
                value = eur_values[-1]

        # If value has no EUR marker in PDF table, infer from remaining numbers.
        if quantity and price and value is None:
            expected = quantity * price
            candidates = [v for v in all_numbers if v > 0 and abs(v - quantity) > 1e-9 and abs(v - price) > 1e-9]
            if candidates:
                value = min(candidates, key=lambda v: abs(v - expected))

        # Strong fallback: compute from quantity and price when no explicit total is detected.
        if quantity and price and value is None:
            value = quantity * price

        # Fallback heuristics when EUR markers are missing entirely
        if price is None and all_numbers:
            non_qty = [v for v in all_numbers if quantity is None or abs(v - quantity) > 1e-9]
            if non_qty:
                # Price tends to be the smaller one, value the larger one
                price = min(non_qty)
                if len(non_qty) > 1:
                    value = max(non_qty)
        
        if quantity is not None and quantity > 0 and (isin or price):
            normalized_name = _normalize_name_by_isin(isin or "", row[1].strip() if len(row) > 1 else "")
            return {
                'category': category or 'Unknown',
                'quantity': quantity,
                'isin': isin or '',
                'name': normalized_name,
                'price': price,
                'value': value,
                'raw_row': ' | '.join([str(c).strip() for c in row])
            }
    except (ValueError, IndexError):
        pass
    
    return None


def _parse_text_for_securities(text: str, securities: List[Dict]) -> None:
    """
    Fallback text parsing method for security extraction.
    Looks for ISIN patterns and associated data.
    """
    # Pattern: ISIN followed by security data
    isin_pattern = r'([A-Z]{2}[A-Z0-9]{9}[0-9])'
    
    lines = text.split('\n')
    current_category = None
    
    for i, line in enumerate(lines):
        # Detect category
        if line.strip() in ['Aktienfonds', 'Rentenwerte', 'Mischfonds', 'Immobilienfonds']:
            current_category = line.strip()
            continue
        
        # Look for ISIN
        isin_match = re.search(isin_pattern, line)
        if isin_match:
            isin = isin_match.group(1)

            quantity = None
            price = None
            value = None

            # Prioritize classic OLB layout where the ISIN is on its own line and
            # the previous/second-previous line contains:
            # "Stück <qty> <name> <price> EUR <value>".
            stueck_line = None
            for offset in (0, 1, 2):
                idx = i - offset
                if idx >= 0:
                    candidate = lines[idx]
                    if re.search(r"St[üu]ck", candidate, re.IGNORECASE) and "EUR" in candidate.upper():
                        stueck_line = candidate
                        break

            # Try to extract name from next line first for two-line layouts,
            # then fallback to previous line.
            name = ''
            if i + 1 < len(lines):
                next_line = lines[i + 1].strip()
                if next_line and 'EUR' in next_line.upper():
                    name = re.sub(r'\s*EUR\s*[0-9][0-9\.,]*\s*$', '', next_line, flags=re.IGNORECASE).strip()
            if not name and i > 0:
                prev_line = lines[i - 1].strip()
                if prev_line and not re.search(r'\d+[\.,]\d+', prev_line[:5]):
                    name = prev_line

            if stueck_line:
                q = _extract_quantity_from_text_fragment(stueck_line)
                if q is not None:
                    quantity = q

                stueck_eur_vals = _extract_eur_numbers_from_fragment(stueck_line)
                if stueck_eur_vals:
                    price = stueck_eur_vals[0]
                    if len(stueck_eur_vals) > 1:
                        value = stueck_eur_vals[1]

                if not name:
                    # Keep text between quantity and first EUR amount as fallback name fragment.
                    n = re.sub(r"^\s*St[üu]ck\s+[0-9][0-9\.,]*\s*", "", stueck_line, flags=re.IGNORECASE)
                    n = re.sub(r"\s+[0-9][0-9\.,]*\s*EUR.*$", "", n, flags=re.IGNORECASE)
                    name = n.strip()

            # Extract price/value from current and nearby lines.
            nearby_idx = range(max(0, i - 2), min(i + 3, len(lines)))
            numeric_context_lines = [lines[j] for j in nearby_idx]

            # --- Specific format detectors (run before generic EUR fallback) ---

            # Format D: Split-ISIN (older Degussa):
            # line i:   "<ISIN> /"           (ISIN alone + slash, WKN cut off)
            # line i+1: "<qty> [<name>] <price>"
            # line i+2: "[<name_cont>] EUR <value>"
            split_isin = re.match(r"^" + re.escape(isin) + r"\s*/\s*$", line.strip())
            if split_isin and i + 1 < len(lines):
                next_parts = lines[i + 1].strip()
                nums_next = re.findall(r"[0-9][0-9\.,]*", next_parts)
                if nums_next:
                    if quantity is None:
                        quantity = _parse_german_number(nums_next[0])
                    if price is None and len(nums_next) >= 2:
                        price = _parse_german_number(nums_next[-1])
                if value is None and i + 2 < len(lines):
                    val_line = lines[i + 2].strip()
                    eur_m = re.search(r"EUR\s+([0-9][0-9\.,]+)", val_line, re.IGNORECASE)
                    if eur_m:
                        value = _parse_german_number(eur_m.group(1))

            # Format C: ISIN-line has "<ISIN> / <WKN> EUR <value>",
            # prev line has "<qty> <name> <price>" (no EUR marker).
            format_c = re.match(
                r"^" + re.escape(isin) + r"\s*/\s*([A-Z0-9]+)\s+EUR\s+([0-9][0-9\.,]+)\s*$",
                line.strip(),
            )
            if format_c:
                if value is None:
                    value = _parse_german_number(format_c.group(2))
                if i > 0:
                    prev = lines[i - 1].strip()
                    nums_prev = re.findall(r"[0-9][0-9\.,]+", prev)
                    if nums_prev:
                        if quantity is None:
                            quantity = _parse_german_number(nums_prev[0])
                        if price is None and len(nums_prev) >= 2:
                            price = _parse_german_number(nums_prev[-1])

            # Format E: ISIN-line has "<name> <ISIN> / <WKN> EUR <value>" (name precedes ISIN),
            # previous line has bare numbers only: "<qty> <price>".
            format_e = re.match(
                r"^.+\s+" + re.escape(isin) + r"\s*/\s*([A-Z0-9]+)\s+EUR\s+([0-9][0-9\.,]+)\s*$",
                line.strip(),
            )
            if format_e:
                if value is None:
                    value = _parse_german_number(format_e.group(2))
                if i > 0 and (quantity is None or price is None):
                    prev = lines[i - 1].strip()
                    # Prev line must consist only of numbers (qty + price, no text)
                    if re.match(r"^[0-9][0-9\.,\s]+$", prev):
                        nums_prev = re.findall(r"[0-9][0-9\.,]+", prev)
                        if nums_prev:
                            if quantity is None:
                                quantity = _parse_german_number(nums_prev[0])
                            if price is None and len(nums_prev) >= 2:
                                price = _parse_german_number(nums_prev[-1])

            # Format B: OLB combined-line layout:
            # "<qty> [<name>] <ISIN> / <WKN> <price>"   (EUR/value on next line)
            olb_block = re.match(
                r"^([0-9][0-9\.,]+)\s+(.*?)\s*" + re.escape(isin) + r"\s*/\s*([A-Z0-9]+)\s+([0-9][0-9\.,]+)\s*$",
                line.strip(),
            )
            if olb_block:
                if quantity is None:
                    quantity = _parse_german_number(olb_block.group(1))
                if not name:
                    name = olb_block.group(2).strip()
                # Always use the price from this line (overrides any earlier generic guess)
                price = _parse_german_number(olb_block.group(4))
                # Value: EUR total on next line
                if value is None and i + 1 < len(lines):
                    next_line_eur = _extract_eur_numbers_from_fragment(lines[i + 1])
                    if next_line_eur:
                        value = next_line_eur[-1]
            else:
                # Fallback: number directly before ISIN on same line (no intervening text)
                line_qty_match = re.search(rf"([0-9][0-9\.,]*)\s+{re.escape(isin)}\b", line)
                if quantity is None and line_qty_match:
                    quantity = _parse_german_number(line_qty_match.group(1))

                line_price_match = re.search(rf"{re.escape(isin)}\s*/\s*[A-Z0-9]+\s+([0-9][0-9\.,]*)", line)
                if price is None and line_price_match:
                    price = _parse_german_number(line_price_match.group(1))

                # No-WKN price: "<ISIN> / <price>" (WKN absent or cut off)
                if price is None:
                    no_wkn = re.search(rf"{re.escape(isin)}\s*/\s*([0-9][0-9\.,]+)", line)
                    if no_wkn:
                        price = _parse_german_number(no_wkn.group(1))

                # First number at start of ISIN line as last-resort quantity
                if quantity is None:
                    first_num = re.match(r"^([0-9][0-9\.,]+)", line.strip())
                    if first_num:
                        quantity = _parse_german_number(first_num.group(1))

                # Value: explicit EUR on next line (for non-olb_block paths)
                if value is None and i + 1 < len(lines):
                    next_line_eur = _extract_eur_numbers_from_fragment(lines[i + 1])
                    if next_line_eur:
                        value = next_line_eur[-1]
                # Also check next line for "EUR <value>" at end
                if value is None and i + 1 < len(lines):
                    combined = lines[i + 1]
                    eur_in_name = re.search(r"EUR\s+([0-9][0-9\.,]+)\s*$", combined, re.IGNORECASE)
                    if eur_in_name:
                        value = _parse_german_number(eur_in_name.group(1))

            # Generic EUR fallback: only used when specific formats above didn't set price/value.
            # Scan closest lines for EUR-tagged numbers.
            if price is None or value is None:
                close_lines = [lines[j] for j in [i, i - 1, i - 2, i + 1] if 0 <= j < len(lines)]
                eur_values = []
                for candidate_line in close_lines:
                    # Skip lines that look like "<WKN> EUR <value>" to avoid WKN being taken as price.
                    # Only accept numbers explicitly tagged with EUR.
                    for m in re.finditer(r"\bEUR\s+([0-9][0-9\.,]+)", candidate_line, re.IGNORECASE):
                        parsed = _parse_german_number(m.group(1))
                        if parsed is not None:
                            eur_values.append(parsed)
                if eur_values and price is None:
                    price = eur_values[0]
                if len(eur_values) > 1 and value is None:
                    value = eur_values[1]


            if quantity is None:
                nearby_lines = [line]
                if i > 0:
                    nearby_lines.append(lines[i - 1])
                if i > 1:
                    nearby_lines.append(lines[i - 2])
                if i + 1 < len(lines):
                    nearby_lines.append(lines[i + 1])

                for candidate_line in nearby_lines:
                    quantity = _extract_quantity_from_text_fragment(candidate_line)
                    if quantity is not None:
                        break

            # Fallback: number directly before ISIN in the same line
            if quantity is None:
                qty_before_isin = re.search(rf"([0-9][0-9\.,]*)\s+{re.escape(isin)}", line)
                if qty_before_isin:
                    quantity = _parse_german_number(qty_before_isin.group(1))

            # 2) If total value is still missing, infer from non-EUR nearby numbers.
            if value is None:
                nearby_numbers = []
                for candidate_line in numeric_context_lines:
                    nearby_numbers.extend(_extract_numbers_from_cell(candidate_line))

                if quantity is not None and price is not None:
                    expected = quantity * price
                    candidates = [
                        v for v in nearby_numbers
                        if v > 0
                        and abs(v - quantity) > 1e-9
                        and abs(v - price) > 1e-9
                    ]
                    if candidates:
                        value = min(candidates, key=lambda v: abs(v - expected))

            # 3) Final fallback: compute total from quantity and price.
            if value is None and quantity is not None and price is not None:
                value = quantity * price
            
            if quantity or price:
                normalized_name = _normalize_name_by_isin(isin, name)
                securities.append({
                    'category': current_category or 'Unknown',
                    'quantity': quantity,
                    'isin': isin,
                    'name': normalized_name,
                    'price': price,
                    'value': value
                })


def export_to_excel(securities: List[Dict], summary_data: Dict, output_path: str) -> None:
    """
    Export securities data to Excel with formatting.
    
    Args:
        securities: List of security dictionaries
        summary_data: Portfolio summary data
        output_path: Path for output Excel file
    """
    # Create workbook
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # Write securities data
        if securities:
            df_securities = pd.DataFrame(securities)
            df_securities.to_excel(writer, sheet_name='Securities', index=False)
        
        # Write summary data
        if summary_data:
            summary_df = pd.DataFrame([
                {
                    'Asset Type': asset_type,
                    'Percentage': data['percentage'],
                    'Value (EUR)': data['value']
                }
                for asset_type, data in summary_data.items()
            ])
            summary_df.to_excel(writer, sheet_name='Summary', index=False)
    
    # Format Excel file
    _format_excel_file(output_path)
    print(f"Excel file saved: {output_path}")


def _format_excel_file(excel_path: str) -> None:
    """Apply formatting to Excel file."""
    wb = openpyxl.load_workbook(excel_path)
    
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        
        # Header formatting
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Format headers
        for cell in ws[1]:
            if cell.value:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
        
        # Build a header-name → column-index map (1-based)
        header_map = {cell.value: cell.column for cell in ws[1] if cell.value}
        date_cols = {header_map[h] for h in ("datum",) if h in header_map}
        numeric_cols = {header_map[h] for h in ("quantity", "price", "value") if h in header_map}

        # Format data cells
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(horizontal='left', vertical='center')

                if cell.value is None:
                    continue

                if cell.column in date_cols:
                    cell.number_format = 'DD.MM.YYYY'
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                elif cell.column in numeric_cols and isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal='right', vertical='center')
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(excel_path)


def parse_folder(folder_path: str, output_xlsx: str = None) -> pd.DataFrame:
    """
    Parse all PDFs in a folder and export results to Excel.
    
    Args:
        folder_path: Path to folder containing PDF files
        output_xlsx: Output Excel file path (default: Depotauszug_Export.xlsx in folder)
        
    Returns:
        DataFrame with extracted data
    """
    folder = Path(folder_path)
    pdf_files = sorted(folder.glob("*.pdf"))
    
    if not pdf_files:
        print(f"No PDF files found in {folder_path}")
        return None
    
    print(f"\n📁 Found {len(pdf_files)} PDF file(s) in {folder.name}")
    print("="*70)
    
    all_securities = []
    all_summary = []
    
    for pdf_file in pdf_files:
        print(f"\n📄 Processing: {pdf_file.name}")
        
        try:
            securities, summary_data = extract_securities_from_pdf(str(pdf_file))
            
            # Add file name and date to each security record
            for sec in securities:
                sec['Datei'] = pdf_file.name
                # 'datum' already set inside extract_securities_from_pdf
            
            all_securities.extend(securities)
            
            # Add to summary
            for asset_type, data in summary_data.items():
                all_summary.append({
                    'Datei': pdf_file.name,
                    'Asset Type': asset_type,
                    'Percentage': data['percentage'],
                    'Value (EUR)': data['value']
                })
            
            print(f"   ✅ Extracted {len(securities)} securities")
            if all_securities:
                best_row = max(
                    securities,
                    key=lambda s: float(s.get('value')) if isinstance(s.get('value'), (int, float)) else float('-inf')
                ) if securities else {}
                top_name = (best_row.get('name') or '')[:50]
                top_value = _format_eur_value(best_row.get('value'))
                print(f"   📊 Top holding: {top_name:50} | €{top_value}")
        
        except Exception as e:
            print(f"   ❌ Error: {e}")
    
    # Create output path
    if output_xlsx is None:
        current_dir = Path(__file__).resolve().parent
        output_xlsx = current_dir / "Depotauszug.xlsx"
    else:
        output_xlsx = Path(output_xlsx)
    
    # Create Excel workbook with multiple sheets
    with pd.ExcelWriter(output_xlsx, engine='openpyxl') as writer:
        if all_securities:
            df_securities = pd.DataFrame(all_securities)
            # Reorder columns for better readability
            cols = ['Datei', 'datum', 'category', 'isin', 'name', 'quantity', 'price', 'value']
            cols = [c for c in cols if c in df_securities.columns]
            df_securities = df_securities[cols]
            df_securities.to_excel(writer, sheet_name='Securities', index=False)
            print(f"\n   📋 Securities sheet: {len(df_securities)} records")
        
        if all_summary:
            df_summary = pd.DataFrame(all_summary)
            df_summary.to_excel(writer, sheet_name='Summary', index=False)
            print(f"   📈 Summary sheet: {len(df_summary)} records")
    
    # Format Excel file
    _format_excel_file(str(output_xlsx))
    
    print(f"\n✅ Exported to: {output_xlsx}")
    print("="*70)
    
    return pd.DataFrame(all_securities) if all_securities else None


def main():
    """Main entry point - process single file or folder."""
    import sys
    
    current_dir = Path(__file__).resolve().parent

    target_dir = Path(r"E:\_NAS\0_Remko\Unterlagen\Banking\_Data\Parse_OLB")
    
    input_path = Path(r"E:\_NAS\0_Remko\Unterlagen\Banking\OLB\DepotAuszug")
    xls_path = Path(target_dir / "Depotauszug.xlsx")
    # Optional CLI override: file or folder path
    if len(sys.argv) > 1:
        input_path = Path(sys.argv[1])

    if not input_path.exists():
        print(f"Error: Path not found: {input_path}")
        return
    
    if input_path.is_file() and input_path.suffix.lower() == '.pdf':
        # Process single file
        print(f"\n📄 Processing single file: {input_path.name}")
        try:
            securities, summary_data = extract_securities_from_pdf(str(input_path))
            output_file = str(xls_path)#input_path.with_stem(input_path.stem + '_Export').with_suffix('.xlsx')
            export_to_excel(securities, summary_data, str(output_file))
            
            print(f"\n✅ Successfully exported to: {output_file}")
            if securities:
                print(f"\nExtracted {len(securities)} securities:")
                for i, sec in enumerate(securities[:5], 1):
                    sec_name = (sec.get('name') or '')[:40]
                    sec_isin = sec.get('isin') or ''
                    sec_value = _format_eur_value(sec.get('value'))
                    print(f"  {i}. {sec_name:40} | ISIN: {sec_isin:12} | €{sec_value}")
                if len(securities) > 5:
                    print(f"  ... and {len(securities) - 5} more")
        except Exception as e:
            print(f"Error: {e}")
            import traceback
            traceback.print_exc()
    
    elif input_path.is_dir():
        # Process all PDFs in folder
        pdf_files = list(input_path.glob('*.pdf'))
        if not pdf_files:
            print(f"No PDF files found in: {input_path}")
            return
        parse_folder(str(input_path),output_xlsx=str(xls_path))
    else:
        print(f"Error: Path not found or invalid: {input_path}")


if __name__ == '__main__':
    main()
